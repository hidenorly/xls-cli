#   Copyright 2022 hidenorly
#
#   Licensed under the Apache License, Version 2.0 (the "License");
#   you may not use this file except in compliance with the License.
#   You may obtain a copy of the License at
#
#       http://www.apache.org/licenses/LICENSE-2.0
#
#   Unless required by applicable law or agreed to in writing, software
#   distributed under the License is distributed on an "AS IS" BASIS,
#   WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
#   See the License for the specific language governing permissions and
#   limitations under the License.

import sys
import os
import argparse
import unicodedata
import openpyxl as xl
import csv

def ljust_jp(value, length, pad = " "):
  count_length = 0
  for char in value.encode().decode('utf8'):
    if ord(char) <= 255:
      count_length += 1
    else:
      count_length += 2
  return value + pad * (length-count_length)

def isRobustNumeric(data):
  if data is not None:
    try:
      float(data)
    except ValueError:
      return False
    return True
  else:
    return False

def getCsvJsonCommon(row):
  result = ""

  for aData in row:
    if result != "":
      result = result + ", "
    if aData is None:
      aData = ""
    if isRobustNumeric(aData):
      result = result + str(aData)
    else:
      result = result + "\"" + str(aData) + "\""

  return result

def getCsv(row):
  result = getCsvJsonCommon(row) + ","
  return result

def getJson(row):
  result = "[ " + getCsvJsonCommon(row) + " ],"
  return result

def dumpSheet(aSheet, args):
  for aRow in aSheet.values:
    if args.json:
      print( "  " + getJson(aRow) )
    elif args.csv:
      print( getCsv(aRow) )

def openBook(fileName):
  result = None

  if os.path.exists( fileName ):
    result = xl.load_workbook( fileName )
  else:
    result = xl.Workbook()

  return result

def openSheet(book, sheetName):
  result = None

  if book:
    for aSheet in book:
      if sheetName == "*" or aSheet.title == sheetName:
        result = aSheet

    if result == None:
      result = book.create_sheet()
      if not sheetName == "*":
        result.title = sheetName

  return result

def getValueArrayFromCells(cells):
  result = []
  for aCol in cells:
    result.append( aCol.value )
  return result

def getLastPosition( sheet ):
  if sheet.max_column==1 and sheet.max_row == 1:
    return 1, 1
  else:
    return 1, sheet.max_row+1

def setCells(targetSheet, startPosX, startPosY, rows, xlSrcCell=True):
  print("hello from " + str(startPosX) + ":" + str(startPosY) )
  if isinstance(targetSheet, xl.worksheet.worksheet.Worksheet):
    y = startPosY
    for aRow in rows:
      x = startPosX
      for aCell in aRow:
        if xlSrcCell:
          targetSheet.cell(row=y, column=x).value = aCell.value
        else:
          targetSheet.cell(row=y, column=x).value = aCell
        x = x + 1
      y = y + 1

def dumpRows(rows):
  for aRow in rows:
    data = getValueArrayFromCells( aRow )
    print( getCsv( data ) )

def getAllSheets(book):
  result = []
  for aSheet in book:
    result.append( aSheet )
  return result

def getSheets(book, targetSheet, enableMerge):
  result = []
  if enableMerge:
    result = getAllSheets( book )
    if len(result) == 0:
     result.append( books.create_sheet() )
  else:
    result.append( openSheet( book, targetSheet ) )
  return result

def getDataFromXlsSheet(aSheet, range, swap):
  resultRows = []

  if range:
    resultRows = aSheet[args.range]
    if swap:
      resultRows = list(map(list, zip(*resultRows)))
  else:
    if swap:
      resultRows = aSheet.columns
    else:
      resultRows = aSheet.rows

  return resultRows

def isXlsBook(filename):
  return filename.endswith(".xlsx") or filename.endswith(".xls") or filename.endswith(".xlsm")


def openCsv( fileName, delimiter ):
  result = []
  if os.path.exists( fileName ):
    file = open( fileName )
    if file:
      reader = csv.reader(file, quoting=csv.QUOTE_MINIMAL, delimiter=delimiter)
      for aRow in reader:
        data = []
        for aCol in aRow:
          aCol = aCol.strip()
          if aCol.startswith("\""):
            aCol = aCol[1:len(aCol)]
          if aCol.endswith("\""):
            aCol = aCol[0:len(aCol)-1]
          data.append( aCol )
        result.append( data )
  return result


def getXlRowColFromXlPos(xlPos):
  xlPos = str( xlPos )
  nLen = len( xlPos )
  nPos = 0
  for i in range( nLen ):
    digit = xlPos[i:i+1]
    if isRobustNumeric( digit ):
      nPos = i
      break
  xlCol = xlPos[0:nPos]
  xlRow = xlPos[nPos:nLen]

  return xlRow, xlCol

def getPositionFromXlLocate(xlPos):
  xlRow, xlCol = getXlRowColFromXlPos( xlPos )
  return int( xlRow ), int( xl.utils.column_index_from_string(xlCol) )

def getRangeFromXlRange(xlPos):
  xlPos = str( xlPos )
  nPos = xlPos.find(":")
  pos = []
  if nPos != -1:
    pos.append( xlPos[0:nPos] )
    pos.append( xlPos[nPos+1:len(xlPos)] )
  result=[]
  for aPos in pos:
    nRow, nCol = getPositionFromXlLocate( aPos )
    result.append( nRow )
    result.append( nCol )

  if len(result) == 4:
    if result[0] >= result[2] and result[1] >= result[3]:
      result[0], result[1], result[2], result[3] = result[2], result[3], result[0], result[1]
    return result[0], result[1], result[2], result[3]

  return 1,1,1,1

def getDataWithRange( rows, xlRange ):
  nStartRow, nStartCol, nEndRow, nEndCol = getRangeFromXlRange( xlRange )
  nColSize = nEndCol - nStartCol + 1
  nRowSize = nEndRow - nStartRow + 1

  nStartRow = nStartRow - 1
  nStartCol = nStartCol - 1

  result = []

  for aRow in range(nRowSize):
    theCol = []
    for aCol in range(nColSize):
      theCol.append( rows[ nStartRow + aRow ][ nStartCol + aCol ] )
    result.append( theCol )

  return result

def getSwappedData( rows ):
  result = [] #list(map(list, zip(*rows)))
  nMaxColSize = 0
  for aRow in rows:
    nSize = len(aRow)
    if nSize > nMaxColSize:
      nMaxColSize = nSize
  for i in range(nMaxColSize):
    result.append( [] )
  for aRow in rows:
    x = 0
    for aCol in aRow:
      result[x].append( aCol )
      x = x + 1

  return result

if __name__=="__main__":
  parser = argparse.ArgumentParser(description='Parse command line options.')
  parser.add_argument('args', nargs='*', help='Specify input.xlsx output.xlsx')
  parser.add_argument('-i', '--inputsheet', action='store', default="*", help='Specfy sheet name e.g. Sheet1')
  parser.add_argument('-o', '--outputsheet', action='store', default="*", help='Specfy sheet name e.g. Sheet1')
  parser.add_argument('-m', '--merge', action='store_true', default=False, help='Specify if you want to merge all of sheets of input book')
  parser.add_argument('-r', '--range', action='store', default=None, help='Specify range e.g. A1:C3 if you want to specify input range')
  parser.add_argument('-s', '--swap', action='store_true', default=False, help='Specify if you want to swap row and column')
  parser.add_argument('-d', '--delimiter', action='store', default=",", help='Specify delimiter for .csv file (default:,)')

  args = parser.parse_args()

  if len(args.args)==2:
    outputBook = openBook( args.args[1] )
    outputSheet = openSheet( outputBook, args.outputsheet )

    if isXlsBook( args.args[0] ):
      inputBook = openBook( args.args[0] )
      inputSheets = getSheets( inputBook, args.inputsheet, args.merge )
      for anInputSheet in inputSheets:
        sourceRows = getDataFromXlsSheet( anInputSheet, args.range, args.swap )
        startPosX, startPosY = getLastPosition( outputSheet )
        setCells( outputSheet, startPosX, startPosY, sourceRows )
    else:
      sourceRows = openCsv( args.args[0], args.delimiter )
      if args.range:
        sourceRows = getDataWithRange( sourceRows, args.range )
      if args.swap:
        sourceRows = getSwappedData( sourceRows )
      startPosX, startPosY = getLastPosition( outputSheet )
      setCells( outputSheet, startPosX, startPosY, sourceRows, False )

    outputBook.save( args.args[1] )
